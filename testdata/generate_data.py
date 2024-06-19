from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
generated_data = Faker()



for i in range(1, 12):
    for j in range(1,4):
        name_list = generated_data.name().split()
        ws.cell(row = i, column = 1 ).value = name_list[0]
        ws.cell(row = i, column = 2 ).value = name_list[1]
        ws.cell(row = i, column = 3 ).value = generated_data.email()
wb.save("testdata\contact_test_data.csv")