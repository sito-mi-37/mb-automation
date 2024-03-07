import csv

from openpyxl import load_workbook


# Define a function to read data from the Excel file
def read_excel_data(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        data.append(row)

    return data


# Define a function to read data from the CSV file
def read_csv_data(file_path):
    with open(file_path, newline='') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)  # Skip the header row if exists
        return [row for row in reader]